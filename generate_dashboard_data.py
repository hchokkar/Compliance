#!/usr/bin/env python3
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

REPORT_DIR = Path('Report')
OUTPUT_PATH = Path('report-data.json')
INDEX_PATH = Path('report-data-index.json')
MONTH_NAME_REGEX = re.compile(r'([A-Za-z]+)\s+(\d{4})')


def parse_report_month(file_path: Path):
    match = MONTH_NAME_REGEX.search(file_path.stem)
    if not match:
        raise ValueError(f'Unable to parse month from filename: {file_path.name}')
    month_name, year = match.groups()
    return datetime.strptime(f'{month_name} {year}', '%B %Y')


def build_report_data(file_path: Path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'Vulnerabilities_All Clients' not in wb.sheetnames:
        raise SystemExit(f'Required sheet not found: Vulnerabilities_All Clients in {file_path.name}')

    ws = wb['Vulnerabilities_All Clients']
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    severity = Counter()
    os_count = Counter()
    state_count = Counter()
    asset_counts = Counter()
    tenant_count = Counter()
    exploitability = Counter()
    month_count = Counter()
    aging_buckets = Counter()
    sla_buckets = Counter()
    top_assets = defaultdict(lambda: Counter())
    plugin_family = Counter()
    critical_assets = Counter()
    unique_locations = Counter()
    asset_types = Counter()
    asset_info = {}
    vulns_by_asset = defaultdict(list)

    all_ages = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        severity[d['Severity'] or 'Unknown'] += 1
        os_value = d['OS'] or 'Unknown'
        os_count[os_value] += 1

        asset_name = d['Assetname'] or d['IP Address'] or 'Unknown'
        asset_counts[asset_name] += 1
        state = d['state'] or 'Unknown'
        state_count[state] += 1
        tenant_count[d['Tenant'] or 'Unknown'] += 1
        exploitability[d['Exploitability'] or 'Unknown'] += 1
        plugin_family[d['Plugin Family'] or 'Unknown'] += 1

        if asset_name not in asset_info:
            asset_info[asset_name] = {
                'ip': d['IP Address'] or 'Unknown',
                'os': d['OS'] or 'Unknown',
                'lastScan': None,
                'status': state,
            }
        else:
            if asset_info[asset_name]['ip'] == 'Unknown' and d['IP Address']:
                asset_info[asset_name]['ip'] = d['IP Address']
            if asset_info[asset_name]['os'] == 'Unknown' and d['OS']:
                asset_info[asset_name]['os'] = d['OS']
            if state != 'Unknown':
                asset_info[asset_name]['status'] = state

        if d['age_in_days'] is not None:
            age = int(d['age_in_days'])
            all_ages.append(age)
            if age <= 30:
                sla_buckets['Within SLA'] += 1
            elif age <= 60:
                sla_buckets['Near SLA'] += 1
            else:
                sla_buckets['Breached SLA'] += 1
            if age <= 30:
                aging_buckets['0-30 Days'] += 1
            elif age <= 60:
                aging_buckets['31-60 Days'] += 1
            elif age <= 90:
                aging_buckets['61-90 Days'] += 1
            elif age <= 180:
                aging_buckets['91-180 Days'] += 1
            elif age <= 365:
                aging_buckets['180+ Days'] += 1
            else:
                aging_buckets['365+ Days'] += 1

        if d['Last Found']:
            try:
                dt = datetime.fromisoformat(str(d['Last Found']))
                month_count[dt.strftime('%Y-%m')] += 1
                current_last = asset_info[asset_name]['lastScan']
                if current_last is None or dt > current_last:
                    asset_info[asset_name]['lastScan'] = dt
            except Exception:
                pass

        if d['Plugin ID'] is not None and d['Severity'] in ('Critical', 'High'):
            critical_assets[asset_name] += 1

        vuln = {
            'pluginId': d['Plugin ID'] or '',
            'pluginName': d['Plugin Name'] or '',
            'severity': d['Severity'] or 'Unknown',
            'protocol': d['Protocol'] or '',
            'port': d['Port'] or '',
            'ip': d['IP Address'] or '',
            'asset': asset_name,
            'pluginOutput': d['Plugin Output'] or '',
            'solution': d['Solution'] or '',
            'firstFound': str(d['First Found']) if d['First Found'] else '',
            'lastFound': str(d['Last Found']) if d['Last Found'] else '',
            'vpr': d['Vulnerability Priority Rating (VPR)'] or '',
            'cve': d['CVE'] or '',
            'description': d['Description'] or '',
            'state': state,
            'age': d['age_in_days'] or '',
            'exploitability': d['Exploitability'] or '',
            'tenant': d['Tenant'] or '',
        }
        vulns_by_asset[asset_name].append(vuln)

        if d['asset.tags']:
            try:
                tags = json.loads(str(d['asset.tags']))
                for tag in tags:
                    if tag.get('category') == 'Server':
                        asset_types[tag.get('value', 'Unknown')] += 1
                    if tag.get('category') == 'Location':
                        unique_locations[tag.get('value', 'Unknown')] += 1
            except json.JSONDecodeError:
                pass

        top_assets[asset_name][d['Severity'] or 'Unknown'] += 1

    severity_total = sum(severity.values())

    asset_rows = []
    for asset, counts in asset_counts.most_common(15):
        info = asset_info.get(asset, {})
        last_scan = info.get('lastScan')
        if isinstance(last_scan, datetime):
            last_scan = last_scan.strftime('%Y-%m-%d')
        asset_rows.append({
            'asset': asset,
            'ip': info.get('ip', 'Unknown'),
            'os': info.get('os', 'Unknown'),
            'status': info.get('status', 'Unknown'),
            'lastScan': last_scan or 'Unknown',
            'totalVulnerabilities': counts,
            'critical': top_assets[asset]['Critical'],
            'high': top_assets[asset]['High'],
            'medium': top_assets[asset]['Medium'],
            'low': top_assets[asset]['Low'],
            'risk': top_assets[asset]['Critical'] * 4 + top_assets[asset]['High'] * 2 + top_assets[asset]['Medium'],
        })

    servers = []
    for asset, counts in asset_counts.most_common(10):
        info = asset_info.get(asset, {})
        last_scan = info.get('lastScan')
        if isinstance(last_scan, datetime):
            last_scan = last_scan.strftime('%Y-%m-%d')
        servers.append({
            'asset': asset,
            'ip': info.get('ip', 'Unknown'),
            'os': info.get('os', 'Unknown'),
            'status': info.get('status', 'Unknown'),
            'lastScan': last_scan or 'Unknown',
            'vulnerabilities': counts,
            'critical': top_assets[asset]['Critical'],
            'high': top_assets[asset]['High'],
            'medium': top_assets[asset]['Medium'],
            'low': top_assets[asset]['Low'],
            'risk': top_assets[asset]['Critical'] * 4 + top_assets[asset]['High'] * 2 + top_assets[asset]['Medium'],
        })

    risk_score = round(
        (severity['Critical'] * 100 + severity['High'] * 60 + severity['Medium'] * 30 + severity['Low'] * 10)
        / max(1, severity_total),
        1,
    )
    compliance_score = round(max(0, 100 - (risk_score * 0.7)), 1)

    report_summary = {
        'totalVulnerabilities': severity_total,
        'totalAssets': len(asset_counts),
        'severity': severity,
        'osDistribution': os_count,
        'state': state_count,
        'tenant': tenant_count,
        'exploitability': exploitability,
        'monthlyTrend': dict(sorted(month_count.items())),
        'sla': sla_buckets,
        'aging': aging_buckets,
        'topServers': servers,
        'assetRows': asset_rows,
        'riskScore': risk_score,
        'complianceScore': compliance_score,
        'averageAge': round(sum(all_ages) / len(all_ages), 1) if all_ages else 0,
        'patchCompliance': round(100 * (exploitability['NOT_AVAILABLE'] / max(1, severity_total)), 1),
        'assetsByOs': dict(os_count.most_common(10)),
        'businessRiskByTenant': dict(tenant_count),
        'assetTypes': dict(asset_types),
        'uniqueLocations': dict(unique_locations),
        'reportMonth': file_path.stem,
    }
    report_details = {
        'reportMonth': file_path.stem,
        'vulnerabilitiesByAsset': {asset: vulns for asset, vulns in vulns_by_asset.items()},
    }
    return report_summary, report_details, parse_report_month(file_path)


if not REPORT_DIR.exists() or not any(REPORT_DIR.glob('*.xlsx')):
    raise SystemExit('No report XLSX files found in Report/')

reports = []
for report_path in sorted(REPORT_DIR.glob('*.xlsx')):
    month_dt = parse_report_month(report_path)
    report_summary, report_details, month_dt = build_report_data(report_path)
    month_key = month_dt.strftime('%Y-%m')
    output_file = Path(f'report-data-{month_key}.json')
    details_file = Path(f'report-data-{month_key}-details.json')

    with output_file.open('w', encoding='utf-8') as f:
        json.dump(report_summary, f, separators=(',', ':'), ensure_ascii=False)

    with details_file.open('w', encoding='utf-8') as f:
        json.dump(report_details, f, separators=(',', ':'), ensure_ascii=False)

    reports.append({
        'month': month_dt.strftime('%B %Y'),
        'key': month_key,
        'file': output_file.name,
        'detailsFile': details_file.name,
        'date': month_dt.isoformat(),
    })

reports.sort(key=lambda item: item['date'])
latest_report = reports[-1]
with OUTPUT_PATH.open('w', encoding='utf-8') as f:
    with Path(latest_report['file']).open(encoding='utf-8') as report_f:
        latest_data = json.load(report_f)
    json.dump(latest_data, f, separators=(',', ':'), ensure_ascii=False)

latest_details_file = Path('report-data-details.json')
with latest_details_file.open('w', encoding='utf-8') as f:
    with Path(latest_report['detailsFile']).open(encoding='utf-8') as details_f:
        latest_details = json.load(details_f)
    json.dump(latest_details, f, separators=(',', ':'), ensure_ascii=False)

index = {
    'reports': reports,
    'defaultReport': latest_report['file'],
}
with INDEX_PATH.open('w', encoding='utf-8') as f:
    json.dump(index, f, indent=2)

print(f'Wrote {OUTPUT_PATH} and index {INDEX_PATH} for {len(reports)} reports.')
